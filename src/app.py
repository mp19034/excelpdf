import openpyxl
from fpdf import FPDF

# Cargar el archivo Excel
excel_dataframe = openpyxl.load_workbook("autoventa.xlsx")

# Seleccionar la hoja activa del archivo Excel
dataframe = excel_dataframe.active

data = []

# Definir el diccionario de códigos y sus umbrales
codes = {
    11974: 8, 12217: 8, 11649: 10, 11664: 10, 11665: 10, 12022: 10,
    11462: 12, 11489: 12, 11654: 12, 11655: 12, 11656: 12, 12011: 12,
    12015: 12, 12041: 12, 12042: 12, 12043: 12, 11531: 14, 11646: 14,
    11647: 14, 11653: 14, 12060: 14, 11977: 14, 12086: 14, 11885: 15,
    11014: 16, 11558: 16, 11560: 16, 11795: 16, 11978: 16, 11993: 16,
    11994: 16, 11995: 16, 11996: 16, 12040: 16, 12090: 16, 12088: 18,
    11867: 18, 11806: 20, 11807: 20, 11808: 20, 11882: 20, 12213: 20,
    12214: 20, 12215: 20, 12216: 20, 11884: 21, 11099: 22, 11860: 25,
    11889: 25, 11064: 30, 11872: 30, 11888: 30, 11890: 30, 11895: 30,
    11038: 32, 11711: 32, 11025: 34, 11098: 34, 11794: 35, 11857: 35,
    11862: 35, 11990: 35, 12044: 35, 12047: 35, 12049: 35, 12050: 35,
    12053: 35, 12054: 35, 12057: 35, 11423: 40, 11463: 40, 11813: 40,
    11864: 40, 11868: 40, 12045: 40, 12218: 40, 12224: 40, 11858: 45,
    11861: 45, 11863: 45, 11883: 45, 11887: 45, 12078: 48, 12123: 48,
    12124: 48, 11211: 50, 11212: 50, 11546: 50, 11874: 50, 11089: 51,
    11091: 51, 11218: 51, 11137: 54, 11139: 54, 11154: 54, 11581: 54,
    11582: 54, 12177: 55, 12127: 57, 11001: 58, 11003: 58, 11007: 58,
    11009: 58, 11027: 58, 11130: 58, 11833: 60, 11834: 60, 11876: 60,
    11880: 60, 11881: 60, 11894: 60, 12007: 60, 12008: 60, 12035: 60,
    12038: 60, 12039: 60, 12046: 60, 12056: 60, 11097: 62, 12126: 63,
    11092: 70, 11094: 70, 11852: 70, 11879: 70, 11886: 70, 12048: 70,
    12051: 70, 12052: 70, 12058: 70, 11793: 72, 11989: 72, 11026: 80,
    11812: 80, 12017: 80, 12018: 80, 12037: 80, 12065: 84, 12066: 84,
    11709: 96, 11710: 96, 11854: 96, 11855: 96, 11856: 96, 12029: 96,
    11873: 100, 12016: 100, 12019: 100, 12003: 110, 12062: 111, 12063: 111,
    12064: 111, 12133: 120, 12131: 122, 12132: 180, 12155: 480, 12156: 480,
    12134: 800
}

# Definir el diccionario de reemplazo para la columna "UMB"
code_replacement = {
    'PAQ': [11001, 11003, 11007, 11009, 11027, 11089, 11091, 11092, 11094, 11130, 11218, 12062, 12063, 12064, 12065, 12066],#ok

    'RT': [11014, 11038, 11462, 11489, 11531, 11558, 11560, 11646, 11647, 11649, 11653, 
           11654, 11655, 11656, 11664, 11665, 11711, 11795, 11806, 11807, 11808, 11974,
           11977, 11978, 11993, 11994, 11995, 11996, 12011, 12015, 12040, 12041, 12042,
           12043, 12060, 12086, 12088, 12090], #ok
    
    'BOL': [11025, 11026, 11064, 11097, 11098, 11099, 11137, 11139, 11154, 11211, 11212,
            11423, 11463, 11546, 11581, 11582, 11709, 11710, 11793, 11794, 11812, 11813,
            11833, 11834, 11852, 11854, 11855, 11856, 11857, 11858, 11860, 11861, 11862,
            11863, 11864, 11867, 11868, 11872, 11873, 11874, 11876, 11879, 11880, 11881,
            11882, 11883, 11884, 11885, 11886, 11887, 11888, 11889, 11890, 11894, 11895,
            11989, 11990, 12003, 12007, 12008, 12016, 12017, 12018, 12019, 12022, 12029,
            12035, 12037, 12038, 12039, 12044, 12045, 12046, 12047, 12048, 12049, 12050,
            12051, 12052, 12053, 12054, 12056, 12057, 12058, 12078, 12123, 12124, 12126,
            12127, 12131, 12132, 12133, 12134, 12177, 12213, 12214, 12215, 12216, 12217, 12218, 12224 ], #ok
    
    'UN': [12155, 12156]
}
#
#

# Iterar para leer los valores de las celdas
#
#
for row in dataframe.iter_rows(min_row=2, max_row=dataframe.max_row-1):
    _row = [row[0].row-1]  # Añadir número de fila
    row[3].value = "No encontrado"
    for cell in row:
        _row.append(cell.value)
    
    # Validar y calcular según las condiciones especificadas
    if len(_row) > 3 and _row[1] in codes:
        threshold = codes[_row[1]]
      
        
        # Reemplazar valor de la columna "UMB"
        for replacement, code_list in code_replacement.items():
            if _row[1] in code_list:
                _row[4] = replacement
                break
        
        if _row[3] >= threshold:
            complete_boxes = _row[3] // threshold
            remaining_units = _row[3] % threshold
            _row.append(complete_boxes)  # Agregar a la quinta columna
            _row.append(remaining_units)  # Agregar a la sexta columna
        else:
            _row.append(0)  # Espacio para la quinta columna
            _row.append(_row[3])  # Agregar a la sexta columna
    else:
        _row.append(0)  # Espacio para la quinta columna
        _row.append(0)  # Espacio para la sexta columna

    data.append(_row)
    #
    #
    #

# Encabezados de las columnas en el PDF
headers = ["#", "Codigo", "Descripcion", "Unidades", "UMB", "Cajas Completas", "Unidades"]

# Texto que deseas agregar encima de la tabla
texto_superior = "Productos Alimenticios Diana, S.A de C.V. \n Distribuidora Santa Ana \n Picking List"

# Texto inferior con tabuladores para alinear Ruta: a la izquierda y Picking: a la derecha
ruta_texto = "Ruta:"
picking_texto = "Picking:"
line = "_________________________________________________________________________________________________________________________________________________________________________________________________________________________________________"
vendedor = "Vendedor:"
carga = "Carga de Mercancia:"
pickig_list = "Picking List:"

# Crear un nuevo PDF en orientación horizontal (landscape)
pdf = FPDF(orientation='L', unit='mm', format='A4')
pdf.add_page()
pdf.set_font("Arial", size=8)

# Calcular el ancho de cada celda para que ajuste al ancho de la página
page_width = pdf.w - 2 * pdf.l_margin  # Ancho de la página menos los márgenes

# Agregar texto encima de los encabezados y la tabla
pdf.set_xy(20, 5)  # Posición XY para el texto superior
pdf.multi_cell(page_width - 20, 8, texto_superior, 0, 'C')
pdf.multi_cell(page_width , 2, line, 0, 'C')

# Calcular la altura del texto superior
texto_superior_height = pdf.get_y()

# Agregar texto inferior con tabuladores para alinear Ruta: a la izquierda y Picking: a la derecha
pdf.set_xy(10, texto_superior_height + 5)  # Posición XY para el texto inferior
pdf.cell(page_width // 2 - 20, 10, ruta_texto, 0, 0, 'L')
pdf.cell(page_width // 2 - 10, 10, picking_texto, 0, 0, 'R')

#Vendedor y carga
pdf.set_xy(10, texto_superior_height + 15)  # Posición XY para el texto inferior
pdf.cell(page_width // 2 - 20, 10, vendedor, 0, 0, 'L')
pdf.cell(page_width // 2 - 10, 10, carga, 0, 0, 'R')

#pickin list
pdf.set_xy(10, texto_superior_height + 25)  # Posición XY para el texto inferior
pdf.cell(page_width // 2 - 20, 10, pickig_list, 0, 0, 'L')

# Agregar la fila de encabezados
pdf.set_xy(pdf.l_margin, pdf.get_y() + 10)  # Posicionar debajo del texto superior
for header in headers:
    pdf.cell(page_width / len(headers), 6, header, 1, 0, 'C')
pdf.ln()

# Agregar las filas de datos
for row_data in data:
    for cell_value in row_data:
        text = str(cell_value) if cell_value is not None else ''
        # Asegurar que el texto esté codificado como UTF-8
        pdf.cell(page_width / len(headers), 6, text.encode('latin-1', 'replace').decode('latin-1'), 1, 0, 'C')
    pdf.ln()

# Guardar el PDF utilizando codificación UTF-8
pdf.output("output.pdf", 'F')
